import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BASE = Path(__file__).resolve().parent
RESULT_JSON = BASE / "kappa_analysis_results.json"
OUTPUT_XLSX = BASE / "kappa_analysis_results.xlsx"


def r4(x):
    return None if x is None else round(float(x), 4)


def pct(x):
    return None if x is None else round(float(x), 4)


def band(ws, row, start_col, end_col, label):
    ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
    cell = ws.cell(row=row, column=start_col, value=label)
    cell.fill = PatternFill("solid", fgColor="1F4E79")
    cell.font = Font(color="FFFFFF", bold=True, size=12)
    cell.alignment = Alignment(horizontal="center")


def write_table(ws, start_row, start_col, headers, rows, table_fill="D9EAF7"):
    thin = Side(style="thin", color="D9E2EC")
    header_fill = PatternFill("solid", fgColor=table_fill)
    for c, header in enumerate(headers, start_col):
        cell = ws.cell(start_row, c, header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    for r, row in enumerate(rows, start_row + 1):
        for c, value in enumerate(row, start_col):
            cell = ws.cell(r, c, value)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)
            if isinstance(value, float):
                cell.number_format = "0.0000"
    return start_row + len(rows)


def auto_width(ws, max_width=34):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for cell in ws[letter]:
            if cell.value is not None:
                width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def main():
    data = json.loads(RESULT_JSON.read_text(encoding="utf-8"))
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    ws["A1"] = "Inter-rater Kappa Analysis"
    ws["A1"].font = Font(size=16, bold=True, color="1F4E79")
    ws["A2"] = "0 = unacceptance, 1 = acceptance; 3 raters: RO, MP, DR"
    ws["A3"] = f"Bootstrap: patient-cluster resampling, n={data['n_bootstrap']}, seed={data['bootstrap_seed']}"

    summary_rows = []
    for model_name, model in data["models"].items():
        ci = data["bootstrap_ci"][model_name]
        pair_ks = [v["kappa"] for v in model["pairwise"].values()]
        summary_rows.append([
            model_name,
            model["n_patients"],
            model["n_items"],
            r4(model["fleiss"]["kappa"]),
            r4(ci["fleiss"][0]),
            r4(ci["fleiss"][1]),
            r4(sum(pair_ks) / len(pair_ks)),
            r4(ci["pair_mean"][0]),
            r4(ci["pair_mean"][1]),
            pct(model["fleiss"]["mean_agreement"]),
            r4(model["fleiss"]["pabak"]),
            pct(model["complete_agreement_rate"]),
            pct(model["rater_acceptance_rate"]["RO"]),
            pct(model["rater_acceptance_rate"]["MP"]),
            pct(model["rater_acceptance_rate"]["DR"]),
        ])
    headers = [
        "Model", "Patients", "Items", "Fleiss kappa", "CI low", "CI high",
        "Mean pairwise kappa", "CI low", "CI high", "Observed agreement",
        "PABAK", "3-rater complete agreement", "RO acceptance", "MP acceptance", "DR acceptance",
    ]
    write_table(ws, 5, 1, headers, summary_rows)

    diff_rows = []
    label_map = {
        "fleiss": "Fleiss kappa",
        "pair_mean": "Mean pairwise kappa",
        "RO-MP": "RO-MP Cohen kappa",
        "RO-DR": "RO-DR Cohen kappa",
        "MP-DR": "MP-DR Cohen kappa",
    }
    for key, diff in data["differences"].items():
        diff_rows.append([
            label_map[key],
            r4(diff["pre_built"]),
            r4(diff["on_site"]),
            r4(diff["difference_on_site_minus_pre"]),
            r4(diff["bootstrap_ci"][0]),
            r4(diff["bootstrap_ci"][1]),
            r4(diff["bootstrap_prob_difference_gt_0"]),
        ])
    write_table(
        ws,
        10,
        1,
        ["Metric", "Pre-built", "On-site trained", "Difference", "CI low", "CI high", "P(diff > 0)"],
        diff_rows,
        "E2F0D9",
    )

    chart = BarChart()
    chart.title = "Model-level Kappa"
    chart.y_axis.title = "Kappa"
    chart.x_axis.title = "Model"
    chart.height = 7
    chart.width = 15
    data_ref = Reference(ws, min_col=4, max_col=7, min_row=5, max_row=7)
    cats = Reference(ws, min_col=1, min_row=6, max_row=7)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend.position = "b"
    ws.add_chart(chart, "P5")

    ws2 = wb.create_sheet("Pairwise")
    ws2.sheet_view.showGridLines = False
    ws2.freeze_panes = "A2"
    pair_rows = []
    for model_name, model in data["models"].items():
        for pair, p in model["pairwise"].items():
            ci = data["bootstrap_ci"][model_name][pair]
            pair_rows.append([
                model_name, pair, p["n"], r4(p["kappa"]), r4(ci[0]), r4(ci[1]),
                pct(p["po"]), r4(p["pabak"]), pct(p["pe"]), p["n00"], p["n01"], p["n10"], p["n11"],
            ])
    write_table(
        ws2,
        1,
        1,
        ["Model", "Rater pair", "N", "Cohen kappa", "CI low", "CI high", "Observed agreement", "PABAK", "Expected agreement", "00", "01", "10", "11"],
        pair_rows,
    )

    ws3 = wb.create_sheet("By question")
    ws3.sheet_view.showGridLines = False
    ws3.freeze_panes = "A2"
    q_rows = []
    for q in data["question_metrics"]:
        q_rows.append([
            q["Model"],
            q["Question"],
            q["n_items"],
            r4(q["Fleiss kappa"]),
            pct(q["Observed agreement"]),
            r4(q["PABAK"]),
            pct(q["Acceptance all ratings"]),
            r4(q["RO-MP kappa"]),
            r4(q["RO-DR kappa"]),
            r4(q["MP-DR kappa"]),
        ])
    write_table(
        ws3,
        1,
        1,
        ["Model", "Question", "N", "Fleiss kappa", "Observed agreement", "PABAK", "Acceptance all ratings", "RO-MP", "RO-DR", "MP-DR"],
        q_rows,
    )

    ws4 = wb.create_sheet("Notes")
    ws4.sheet_view.showGridLines = False
    notes = [
        ["Interpretation", "Kappa is chance-corrected agreement. Common descriptive bands: <0 none/slight, 0.21-0.40 fair, 0.41-0.60 moderate, 0.61-0.80 substantial, >0.80 almost perfect."],
        ["Unit of analysis", "Each patient-question judgment was treated as one binary item. There are 30 patients and 17 questions per model, so N=510 items per model."],
        ["Fleiss kappa", "Used for overall agreement among RO, MP, and DR."],
        ["Cohen kappa", "Used for each rater pair: RO-MP, RO-DR, MP-DR."],
        ["PABAK", "Prevalence-adjusted bias-adjusted kappa, calculated as 2 x observed agreement - 1."],
        ["Bootstrap CI", "95% percentile intervals from patient-cluster bootstrap resampling, preserving all 17 questions within each sampled patient."],
        ["Difference", "On-site trained minus Pre-built. Negative values mean lower chance-corrected agreement for the on-site trained model."],
    ]
    write_table(ws4, 1, 1, ["Topic", "Note"], notes, "FCE4D6")
    ws4.column_dimensions["A"].width = 24
    ws4.column_dimensions["B"].width = 120
    ws4["B1"].alignment = Alignment(wrap_text=True)

    for sheet in wb.worksheets:
        auto_width(sheet)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    horizontal=cell.alignment.horizontal,
                    vertical="center",
                    wrap_text=True,
                )
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["P"].width = 18
    ws.column_dimensions["Q"].width = 18
    ws.column_dimensions["R"].width = 18
    ws.column_dimensions["S"].width = 18
    wb.save(OUTPUT_XLSX)
    print(OUTPUT_XLSX)


if __name__ == "__main__":
    main()
